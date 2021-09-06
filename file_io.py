import os
import datetime
import mne
import pyedflib
import numpy as np
import constants
from scipy import io

from thrift import TSerialization
from thrift_model.TSignalDesc.ttypes import SignalLabel, DimensionType, TransducerType, EyeStatus
from thrift_model.TSignalSession.ttypes import TSignalSession
from openpyxl import Workbook

from pathlib import Path
from config.cfg import cfg


def build(mPmax, mPmin, mDmax, mDmin):
    return (mPmax - mPmin) / float(mDmax - mDmin)


# fixme : it takes too long : 11
def read_signal(signal, mUnitInDigit):
    # fixme : it takes too long : 8
    def readDigValue24Bits(src, pos):
        def read24BitsByteL(src, pos):
            return ((src[pos + 2] & 0x000000FF) << 16) | ((src[pos + 1] & 0x000000FF) << 8) | ((src[pos] & 0x000000FF))

        read24BitsByteL.__doc__ = "Default function! Converting 24 bits number to 3 bytes number"

        dValue = read24BitsByteL(src, pos)
        if (dValue & 0x800000) > 1:
            dValue = dValue - (1 << 24)
        else:
            dValue &= 0xffffff
        dValue *= -1
        return dValue

    readDigValue24Bits.__doc__ = "Default function! Converts physical values to digital value"

    sizeOfDigit = 3
    nDigitValues = int(len(signal) / sizeOfDigit)
    pValues = []
    for i in range(0, nDigitValues):
        dValue = readDigValue24Bits(signal, sizeOfDigit * i)
        pValues.append(dValue * mUnitInDigit)
    return pValues


def read_file(info, crt_file_name):
    if crt_file_name.endswith('.edf'):
        input_data, fs, ch_names, data_types, birthday = read_edf_file(crt_file_name)
        if birthday is not None:
            info['subject']['birthday'] = birthday
        crt_ytdf_ref_file = None
        crt_edf_data = {'data': input_data,
                        'fs': fs,
                        'ch_names': ch_names,
                        'data_types': data_types}
    elif crt_file_name.endswith('.ytdf'):
        crt_edf_data = None
        crt_ytdf_ref_file = os.path.join(crt_file_name)
    elif crt_file_name.endswith('.mat'):
        input_data, fs, ch_names, data_types = read_mat_file(os.path.join(crt_file_name), 'baseline')
        crt_ytdf_ref_file = None
        crt_edf_data = {'data': input_data,
                        'fs': fs,
                        'ch_names': ch_names,
                        'data_types': data_types}
    else:
        crt_ytdf_ref_file = None
        crt_edf_data = None

    return crt_ytdf_ref_file, crt_edf_data


def read_edf_file(path_file):
    reader = pyedflib.EdfReader(path_file)

    fs = reader.getSampleFrequencies()[0]
    ch_label_tmp = reader.getSignalLabels()
    ch_label_list = []
    for ch in ch_label_tmp:
        ch_label_list.append(ch.replace('EEG', '').replace(' ', '').upper())

    sig_bufs = np.zeros([ch_label_list.__len__(), reader.getNSamples()[0]])

    for ch in range(sig_bufs.__len__()):
        sig_bufs[ch] = reader.readSignal(ch)

    if 'A1' in ch_label_list and 'A2' in ch_label_list:
        a1_idx = ch_label_list.index('A1')
        a2_idx = ch_label_list.index('A2')
        sig_bufs -= (sig_bufs[a1_idx] + sig_bufs[a2_idx]) / 2
        if a2_idx > a1_idx:
            sig_bufs = np.delete(sig_bufs, a2_idx, 0)
            sig_bufs = np.delete(sig_bufs, a1_idx, 0)
            ch_label_list.remove('A2')
            ch_label_list.remove('A1')
    elif 'A1' in ch_label_list:
        a1_idx = ch_label_list.index('A1')
        sig_bufs -= sig_bufs[a1_idx]
        sig_bufs = np.delete(sig_bufs, a1_idx, 0)
        ch_label_list.remove('A1')
    elif 'A2' in ch_label_list:
        a2_idx = ch_label_list.index('A2')
        sig_bufs -= sig_bufs[a2_idx]
        sig_bufs = np.delete(sig_bufs, a2_idx, 0)
        ch_label_list.remove('A2')

    if 'CZ' not in ch_label_list:
        cz_data = np.expand_dims(np.mean(sig_bufs, 0), 0)
        sig_bufs = np.concatenate([sig_bufs, cz_data], 0)
        ch_label_list.append('CZ')

    input_data = np.zeros([1, 19, reader.getNSamples()[0]])
    # input_data = np.zeros([1, sig_bufs.shape[0], sig_bufs.shape[1]])
    new_ch_label_list = []
    new_data_type_list = []

    for ch in range(constants.YBRAIN_CH_LIST.__len__()):
        input_data[0, ch] = sig_bufs[ch_label_list.index(constants.YBRAIN_CH_LIST[ch].upper())]
        new_ch_label_list.append(constants.YBRAIN_CH_LIST[ch])
        new_data_type_list.append('eeg')

    try:
        birthday = datetime.datetime.strptime(str(reader.patient).split(' ')[2], '%d-%b-%Y')
        birthday = birthday.date().strftime('%Y%m%d')
        return input_data, fs, new_ch_label_list, new_data_type_list, birthday
    except:
        return input_data, fs, new_ch_label_list, new_data_type_list, None


def read_mat_file(path_file, type='baseline'):
    mat_file = io.loadmat(path_file)
    data = mat_file[type]
    ch_list = ['FP1', 'FP2', 'F3', 'F4', 'C3', 'C4', 'P3', 'P4', 'O1', 'O2', 'F7', 'F8', 'T3', 'T4', 'T5', 'T6', 'FZ',
               'CZ', 'PZ']

    input_data = np.zeros([1, 19, data.shape[1]])
    new_ch_label_list = []
    new_data_type_list = []
    for ch in range(constants.YBRAIN_CH_LIST.__len__()):
        input_data[0, ch] = data[ch_list.index(constants.YBRAIN_CH_LIST[ch].upper())]
        new_ch_label_list.append(constants.YBRAIN_CH_LIST[ch])
        new_data_type_list.append('eeg')

    return input_data, 500, new_ch_label_list, new_data_type_list


def read_bin_data(ytdf_dir, specs=None):
    with open(ytdf_dir, "rb") as ytdf_file:
        blog2 = ytdf_file.read()

    blog = TSignalSession()
    tmp = TSerialization.deserialize(blog, blog2)

    tmp.validate()
    FirstElem = True
    channel_names = []
    channel_types = []
    montage = 'standard_1020'
    signal_per_ch_limits = dict()
    bad_channels, channels_of_interest = [], []
    for signal in tmp.signals:

        if signal.adcBits == 24:
            mUnitInDigit = build(signal.phyMax, signal.phyMin, signal.digMax, signal.digMin)
            exResults = np.array(read_signal(signal.data, mUnitInDigit))
            if FirstElem:
                FirstElem = False
                SignalsProcessed = np.array([exResults])
            else:
                SignalsProcessed = np.concatenate((SignalsProcessed, np.array([exResults])))

            dimension_temp = DimensionType._VALUES_TO_NAMES[signal.dimension]
            if not dimension_temp in ["nV", "mV", "V"]:
                dimension_temp = "uV"
            signal_per_ch_limits[SignalLabel._VALUES_TO_NAMES[signal.label]] = {"phyMax": signal.phyMax,
                                                                                "phyMin": signal.phyMin,
                                                                                "digMax": signal.digMax,
                                                                                "digMin": signal.digMin,
                                                                                "dimension": dimension_temp}
            del dimension_temp
            if SignalLabel._VALUES_TO_NAMES[signal.label] == "FP1":
                ch_names_from_mne = "Fp1"
            elif SignalLabel._VALUES_TO_NAMES[signal.label] == "FP2":
                ch_names_from_mne = "Fp2"
            else:
                ch_names_from_mne = SignalLabel._VALUES_TO_NAMES[signal.label]
            channel_names.append(ch_names_from_mne)
            channel_types.append('eeg')
            if specs:
                for i in range(len(specs)):
                    if specs[i] == "FP1":
                        specs[i] = "Fp1"
                    elif specs[i] == "FP2":
                        specs[i] = "Fp2"
                if not ch_names_from_mne in specs:
                    channels_of_interest.append(ch_names_from_mne)
            if signal.badChannelSignal:
                bad_channels.append(ch_names_from_mne)
    if tmp.artifactSegmentRejection:
        sig_annotations = tmp.artifactAnnotations
    else:
        sig_annotations = tmp.sigAnnotations

    if channels_of_interest == []:
        channels_of_interest = None
    eeg_info = {"eegSamplingRate": tmp.eegSamplingRate, "eegSamples": tmp.eegSamples, "recStartTime": tmp.recStartTime,
                "recStopTime": tmp.recStopTime, "eyeStatus": EyeStatus._VALUES_TO_NAMES[tmp.eyeStatus],
                "name": tmp.name, "description": tmp.description,
                "impedanceOn": tmp.impedanceOn, 'signal_per_ch_limits': signal_per_ch_limits,
                "transducer": TransducerType._VALUES_TO_NAMES[tmp.transducer],
                "subject": tmp.subject, "annotations": sig_annotations, "channels_of_interest": channels_of_interest}
    fs = eeg_info['eegSamplingRate']
    info = mne.create_info(channel_names, fs, channel_types, montage)
    del channel_names, channel_types, montage
    info['description'] = 'Ybrain dataset'

    try:
        if SignalsProcessed.shape.__len__() == 2:
            custom_raw = mne.io.RawArray(SignalsProcessed, info)
        else:
            custom_raw = mne.io.RawArray(SignalsProcessed[0], info)
        del SignalsProcessed, info
    except:
        raise RuntimeError('Too large data!!! Need optimization!')

    return custom_raw, eeg_info, bad_channels


def custom_raw_from_edf(data_eeg, fs, channel_names, channel_types, montage='standard_1020'):
    info = mne.create_info(channel_names, fs, channel_types, montage)
    custom_raw = mne.io.RawArray(data_eeg[0], info)
    return custom_raw


def make_mne_raw(ytdf_dir, data_edf=None, specs=None):
    if ytdf_dir:
        mne_raw, eeg_info, bad_channels = read_bin_data(ytdf_dir, specs)
    elif data_edf:
        data_eeg = data_edf['data']
        fs = data_edf['fs']
        ch_names = data_edf['ch_names']
        data_types = data_edf['data_types']
        mne_raw = custom_raw_from_edf(data_eeg, fs, ch_names, data_types)
        eeg_info = None
        bad_channels = []
    else:
        raise RuntimeError('Source not exists.')

    return mne_raw, eeg_info, bad_channels


def write_excel(cdf_dict, raw_dict, f, psd, request_id):
    YB_CH_NAME_LIST = ['', 'Fp1', 'F7', 'T3', 'T5', 'T6', 'T4', 'F8', 'Fp2', 'F4', 'C4', 'P4', 'O2', 'Pz', 'Fz', 'Cz', 'O1', 'P3', 'C3', 'F3']
    write_wb = Workbook()

    if raw_dict is not None:
        for feature in ['abs_power', 'rel_power', 'rat_power']:
            write_ws = write_wb.create_sheet('%s_raw' % (feature))
            write_ws.append(YB_CH_NAME_LIST)
            for band in raw_dict[feature].keys():
                power = raw_dict[feature][band]
                write_ws.append([band] + list(power))  # list

        if raw_dict['alpha_peak'] is not None:
            write_ws = write_wb.create_sheet('alpha peak')
            write_ws.append(YB_CH_NAME_LIST)
            write_ws.append(['alpha peak frequency'] + list(raw_dict['alpha_peak']))
            write_ws.append(['alpha peak power'] + list(raw_dict['alpha_peak_power']))

        if raw_dict['coh'] is not None:
            for band in raw_dict['coh'].keys():
                cur_coh = raw_dict['coh'][band]
                write_ws = write_wb.create_sheet('coherence_%s' % band)
                write_ws.append(YB_CH_NAME_LIST)
                for ch in range(19):
                    write_ws.append([YB_CH_NAME_LIST[ch + 1]] + list(cur_coh[ch]))

    if cdf_dict is not None:
        for feature in ['abs_power', 'rel_power', 'rat_power']:
            write_ws = write_wb.create_sheet('%s_cdf' % (feature))
            write_ws.append(YB_CH_NAME_LIST)
            for band in cdf_dict[feature].keys():
                power = cdf_dict[feature][band]
                write_ws.append([band] + list(power))  # list

        if cdf_dict['stress'] is not None:
            write_ws = write_wb.create_sheet('summary_score')
            summary_feature_list = ['stress', 'cognition', 'concentration', 'use_of_brain', 'memory_operate', 'info_amount', 'info_speed', 'info_complex']
            write_ws.append(summary_feature_list)
            summary_value_list = []
            for feature in summary_feature_list:
                summary_value_list.append(cdf_dict[feature])
            write_ws.append(summary_value_list)

            write_ws = write_wb.create_sheet('psd')
            write_ws.append(YB_CH_NAME_LIST)
            for i in range(np.where(f <= 50)[0].shape[0]):
                write_ws.append(['%.4fHz' % f[i]] + list(psd[:, i]))

    write_wb.remove_sheet(write_wb['Sheet'])
    path_csv = Path(cfg.OUT_DIR, request_id)
    path_csv.mkdir(exist_ok=True, parents=True)
    write_wb.save(path_csv / Path('feature_val.xlsx'))
